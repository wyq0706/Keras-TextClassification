from openpyxl import load_workbook


def readData():
    workbook = load_workbook(u'/home/liuaohan/sentence label.xlsx')
    booksheet = workbook.active

    calc = {"has_shape": {"all": 0,
                          "is_none": 0,
                          "is_chip": 0,
                          "is_dot": 0,
                          "is_ban": 0,
                          "is_tuan": 0,
                          "is_tiao": 0,
                          "is_xian": 0,
                          "is_jiejie": 0,
                          "is_dai": 0,
                          "is_dot_chip":0,
                          "is_ban_dian":0,
                          "is_ban_chip":0,
                          "is_tiao_chip":0},
            "has_chip_shape": 0,
            "has_dot_shape": 0,
            "has_density": {"all": 0,
                            "is_none": 0,
                            "normal": 0},
            "has_density_low": 0,
            "has_density_high": 0,
            "has_density_both": 0,
            "has_grown": {
                "duofa": 0,
                "sanfa": 0,
                "both": 0,
                "none": 0,
            },
            "all_num": 0}

    for index, row in enumerate(booksheet.rows):
        line = [col.value for col in row]
        if index == 0:
            continue
        if line[2] is not None and len(line[2].strip()) != 0:
            calc["all_num"] += 1
            sentence_str = line[1].strip()
            label_str = [i for i in line[2].strip().split(" ") if i != ""]
            if len(label_str) != 1 or label_str[0] != "无":
                if sentence_str.find("多发") != -1:
                    calc["has_grown"]["duofa"] += 1
                elif sentence_str.find("散发") != -1:
                    calc["has_grown"]["sanfa"] += 1
                else:
                    print("grow:", label_str, sentence_str)
                    calc["has_grown"]["none"] += 1
            if sentence_str.find("密度") != -1:
                calc["has_density"]["all"] += 1
                if sentence_str.find("低密度") != -1 or sentence_str.find("密度减低") != -1 or sentence_str.find(
                        "密度略减低") != -1 or sentence_str.find("密度变低") != -1:
                    calc["has_density_low"] += 1
                elif sentence_str.find("高密度") != -1 or sentence_str.find("密度增高") != -1:
                    calc["has_density_high"] += 1
                elif sentence_str.find("未见异常密度影") != -1 or sentence_str.find("未见明显异常密度影") != -1:
                    calc["has_density"]["normal"] += 1
                elif sentence_str.find("混杂密度影") != -1 or sentence_str.find("密度不均匀") != -1 or sentence_str.find(
                        "高低混杂密度") != -1:
                    calc["has_density_both"] += 1
                elif len(label_str) != 1 or label_str[0] != "无":
                    # else:
                    print("density: ", label_str, sentence_str)
                # if sentence_str.find("低") != -1 and sentence_str.find("高") != -1:
                #     # print(label_str, sentence_str)
                #     pass
                if len(label_str) == 1 and label_str[0] == "无":
                    calc["has_density"]["is_none"] += 1
            if sentence_str.find("状") != -1 and (len(label_str) != 1 or label_str[0] != "无"):
                calc["has_shape"]["all"] += 1
                if sentence_str.find("点片状") != -1:
                    calc["has_chip_shape"] += 1
                    calc["has_shape"]["is_dot_chip"] += 1
                elif sentence_str.find("条片状") != -1:
                    calc["has_chip_shape"] += 1
                    calc["has_shape"]["is_tiao_chip"] += 1
                elif sentence_str.find("斑片状") != -1:
                    calc["has_chip_shape"] += 1
                    calc["has_shape"]["is_ban_chip"] += 1
                elif sentence_str.find("片状") != -1:
                    calc["has_chip_shape"] += 1
                    calc["has_shape"]["is_chip"] += 1
                # elif sentence_str.find("斑点状") != -1:
                #     calc["has_chip_shape"] += 1
                #     calc["has_shape"]["is_ban_dian"] += 1
                elif sentence_str.find("点状") != -1:
                    calc["has_dot_shape"] += 1
                    calc["has_shape"]["is_dot"] += 1
                elif sentence_str.find("斑状") != -1:
                    calc["has_shape"]["is_dot"] += 1
                # elif sentence_str.find("团状") != -1:
                #     calc["has_shape"]["is_tuan"] += 1
                elif sentence_str.find("条状") != -1:
                    calc["has_shape"]["is_tiao"] += 1
                elif sentence_str.find("线状") != -1:
                    calc["has_shape"]["is_tiao"] += 1
                # elif sentence_str.find("结节状") != -1:
                #     calc["has_shape"]["is_jiejie"] += 1
                elif sentence_str.find("带状") != -1:
                    calc["has_shape"]["is_tiao"] += 1
                elif len(label_str) != 1 or label_str[0] != "无":
                    print("shape: ", label_str, sentence_str)
                if len(label_str) == 1 and label_str[0] == "无":
                    calc["has_shape"]["is_none"] += 1

    print(calc)
    sum_=0
    for index,item in enumerate(calc["has_shape"]):
        if index<=1:
            continue
        print(item)
        print(calc["has_shape"][item]/(calc["has_shape"]["all"]-calc["has_shape"]["is_none"]))
        sum_+=calc["has_shape"][item]/(calc["has_shape"]["all"]-calc["has_shape"]["is_none"])
    print(sum_)


def writeData():
    workbook = load_workbook(u'/home/liuaohan/sentence label.xlsx')
    booksheet = workbook.active
    sum_list = []
    f_all = open('./data/all.csv', 'w', encoding='utf-8')
    for index, row in enumerate(booksheet.rows):
        line = [col.value for col in row]
        if index == 0:
            continue
        if line[2] is not None and len(line[2].strip()) != 0:
            label_str = ",".join(i for i in line[2].strip().split(" ") if i != "")
            add_str = "\n" + "|,|".join([label_str, "".join(line[1].strip().split("\n"))])
            sum_list.append(add_str)
    f_all.writelines(sum_list)
    f_all.close()


if __name__ == "__main__":
    readData()
    # writeData()
